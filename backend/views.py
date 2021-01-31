from django.shortcuts import render
from django.http import HttpResponse
from django import forms
import csv
import codecs
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import numpy as np
import math
from datetime import *
from itertools import chain
import pickle
import os


def reconcile_transfers(data,matches=list()):
	if len(data) == 0:
		return matches

	# matching algorithm
	match_with = data[0]
	matched = list()
	for trxn in data:
		if trxn[5] != match_with[5] and trxn[8]==match_with[8]*-1 and trxn[6]==match_with[6]:
			matched.append(trxn)
	if len(matched)==0:
		# try again with the next day
		for trxn in data:
			if trxn[5] != match_with[5] and trxn[8]==match_with[8]*-1 and abs((trxn[6])-match_with[6])<3:
				matched.append(trxn)

	if len(matched)==1:
		index = matched[0][0]
		matches.extend([match_with[0],index])
		data = [row for row in data if row[0] != index]
	return reconcile_transfers(data[1:],matches)


def date_case_a(date):
	try:
		return datetime.strptime(date,"%Y-%m-%d").toordinal()
	except:
		return -1

def date_case_b(date):
	try:
		return datetime.strptime(date,"%d/%m/%y").toordinal()
	except:
		return -1

def date_handle(date):
	new_date = date_case_a(date)
	if(new_date > 0):
		return new_date
	new_date = date_case_b(date)
	if(new_date > 0):
		return new_date
	raise Exception(date,"DATE FORMAT NOT SUPPORTED")


def triage(data,header):
	# index data
	data = [[i]+row for i,row in enumerate(data)]
	header = ['key']+header

	# split into transfers and non transfers
	transfer_categories = ["External Transfers","Internal Transfer","Credit Card Repayments"]
	transfers = [row for row in data if row[10] in transfer_categories]
	transfers = [i[:6]+[date_handle(i[6])]+[i[7]]+[float(i[8])]+i[9:] for i in transfers]
	non_transfers = [row for row in data if row[10] not in transfer_categories]
	# match transfers
	transfer_matches = reconcile_transfers(transfers)
	#transfer_numbers_flat = [index for match in transfer_matches for index in match]
	unmatched_transfers = [row for row in transfers if row[0] not in transfer_matches]
	transfers = [row for row in transfers if row[0] in transfer_matches]
	# sort according to transfer_matches indexes
	transfers = list(chain(*[[row for row in transfers if row[0] == i] for i in transfer_matches]))
	# split categorised and non_categorised
	categorised = [row for row in non_transfers if len(row[10])>0]
	non_categorised = [row for row in non_transfers if len(row[10])<1]

	# put the dates back from ordinal to iso in transfers
	transfers = [row[:6]+[datetime.fromordinal(row[6]).date().isoformat()]+row[7:] for row in transfers]
	unmatched_transfers = [row[:6]+[datetime.fromordinal(row[6]).date().isoformat()]+row[7:] for row in unmatched_transfers]

	return {
		'transfers': [header]+transfers,
		'unmatched_transfers': [header]+unmatched_transfers,
		'categorised': [header]+categorised,
		'non_categorised': [[head for head in header]]+non_categorised
	}

def classify(data,model):
	data[0][11] = "Confidence"
	for trxn in data[1:]:
		categories = [category for category in model]
		words = trxn[7].split(' ')
		x = np.array([[model[category][word] if word in model[category] else 0 for category in categories] for word in words])
		xsums = np.sum(x,axis=0)
		trxn[10] = categories[np.argmax(np.sum(x,axis=0))]
		trxn[11] = 0
		if sum(xsums) != 0:
			trxn[11] = max(xsums)/sum(xsums)
	return data

def dump_spreadsheet(data):
	workbook = Workbook()
	sheet_one = workbook.active
	sheet_one.title = "uncategorised"
	for row in data['non_categorised']:
		sheet_one.append(row)

	sheet_two = workbook.create_sheet(title="categorised")

	for row in data['categorised']:
		sheet_two.append(row)

	sheet_three = workbook.create_sheet(title="matched transfers")
	for row in data['transfers']:
		sheet_three.append(row)

	sheet_four = workbook.create_sheet(title="unmatched transfers")
	for row in data['unmatched_transfers']:
		sheet_four.append(row)

	return workbook


class Form(forms.Form):
	file = forms.FileField()

# Create your views here.
def index(request):
	return render(request,"index.html")

def run(request):
	submitted = Form(request.POST,request.FILES)
	if submitted.is_valid():
		f = submitted.cleaned_data['file']
		data = [row for row in csv.reader(codecs.iterdecode(f, 'utf-8'),delimiter=',')]


		response = HttpResponse(content_type='text/csv')

		module_dir = os.path.dirname(__file__)  # get current directory
		file_path = os.path.join(module_dir, 'model.pickle')
		model = pickle.load(open(file_path,'rb'))
		data = triage(data[1:],data[0])
		data['non_categorised'] = classify(data['non_categorised'],model)
		workbook = dump_spreadsheet(data)

		response = HttpResponse(save_virtual_workbook(workbook), content_type='application/vnd.ms-excel')
		response['Content-Disposition'] = 'attachment; filename="living_expenses.xlsx"'

		print("living expenses spreadsheet generated")

		return response

	return HttpResponse(status=200)
